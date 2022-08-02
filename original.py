# -*- coding: utf8 -*-
# Модель:
# "Оценка влияния показателей социально-экономического развития на валовой территориальный продукт г.Казани"
#
# Исходные данные:
# 1. Валовой территориальный продукт, млрд. руб. (VTP_t)
# 2. Добавленная стоимость, млрд. руб. (DST_t)
# 3. Добавленная стоимость в разрезе видов экономической деятельности, млрд. руб. (DTSotrasli):
#       3.1 брабатывающие производства
#       3.2 строительство
#       3.3 транспортировка и хранение
#       3.4 оптовая и розничная торговля
#       3.5 деятельность в области информатизации и связи
# 4. темп роста валового территориального продукта в основных ценах, процент VTPplus
# 5. темп роста добавленной стоимости в разрезе видов (см. выше) экономической деятельности, процент (DSTplusOtrasli)
#
# Результат:
# 1. Влияние прироста добавленной стоимости на ВТП в определенные периоды моделирования, процент  (deltaVDST)
# 2. Влияние прироста добавленной стоимости в разрезе видов экономической деятельности на ВТП в определенные
#    периоды моделирования, процент (deltaVDST)
import pandas as pd
from typing import Dict, List
import openpyxl
from openpyxl.styles import Alignment
import plotly.express as px
import doctest 

document_inner = pd.read_excel("C:/Users/pc/.vscode/gitwork/histograms/doctest_files/data.xlsx")
years_count = 4  # Количество лет для прогноза

CORRECT_TITLES = ['ВТП Казани, тыс. руб.', 'Добавленная стоимость, тыс. руб.',
                  'ДС Обрабатывающие производства, тыс. руб.', 'ДС Строительство, тыс. руб.',
                  'ДС Транспортировка и хранение, тыс. руб.', 'ДС Оптовая и розничная торговля, тыс. руб.',
                  'ДС Деятельность в области информатизации и связи, тыс. руб.']


def check_params(data_frame_object: pd.DataFrame) -> bool:
    '''
        Функция проверяет параметры в таблице на ошибки
        На вход функции подается таблица в виде DataFrame объекта
        На выход функции подается значения типа boolean для пояснения о наличии или отсутствии ошибки

        Returns:
            True - first collumn in data_frame_object contains 
                   all items from CORRECT_TITLES
            False - otherwise

        >>> df = pd.read_excel('wrong1.xlsx')
        >>> check_params(df)
        С параметром на 7 строке ошибка
        False

        >>> df = pd.read_excel('wrong3.xlsx')
        >>> check_params(df)
        С параметром на 4 строке ошибка
        False

        >>> df = pd.read_excel('correct1.xlsx')
        >>> check_params(df)
        True

        >>> df = pd.read_excel('correct2.xlsx')
        >>> check_params(df)
        True
    '''
    found_mistake = False

    for index_tt, title in enumerate(data_frame_object.values):
        if title[0] not in CORRECT_TITLES:
            print(f'С параметром на {index_tt + 2} строке ошибка')
            found_mistake = True

    return not found_mistake


def is_empty(data_frame_object: pd.DataFrame) -> bool:
    '''
        Функция проверяет параметры таблицы на пустоту 
        На вход функции подается таблица в виде DataFrame объекта
        На выход функции подается значения типа boolean для пояснения о наличии или отсутствии ошибки

        Returns:
            True - in data_frame_object are not any empty cells
            False - otherwise

        Function checks filling of DF object, if there
        are missing data, function returns True and prints
        where this mistake are. Else, if there aren't
        any problems, function returns False.

        >>> df = pd.read_excel('correct1.xlsx')
        >>> is_empty(df)
        False

        >>> df = pd.read_excel('wrong1.xlsx')
        >>> is_empty(df)
        Ошибка в значении  ДС Транспортировка и хранение, тыс. руб. за  Y-3 год
        Ошибка в значении  ДС Транспортировка и хранение, тыс. руб. за  Y-2 год
        True

        >>> df = pd.read_excel('wrong2.xlsx')
        >>> is_empty(df)
        Ошибка в значении  ДС Обрабатывающие производства, тыс. руб. за  Y-1 год
        Ошибка в значении  ДС Транспортировка и хранение, тыс. руб. за  Y-2 год
        Ошибка в значении  ДС Оптовая и розничная торговля, тыс. руб. за  Y-3 год
        True

    '''
    found_empty = False

    for string in data_frame_object.values:
        for i, element in enumerate(string[1:]):
            if pd.isna(element):
                found_empty = True
                print(
                    f"Ошибка в значении  {string[0]} за  {data_frame_object.columns[i + 1]} год")

    return found_empty


def get_temp_rosta(DS_otrasli: Dict[str, List[float]]) -> Dict[str, float]:
    '''
        Возвращает Темп роста добавленной стоимости в разрезе видов (см. выше) экономической деятельности, процент 
        На вход функции подается словарь со значениями из таблицы 
        На выход функции подается словарь с коэффициентами показателя темпа роста

        >>> data =  {'ВТП Казани, тыс. руб.': [674739900.0, 748722900.0, 773031100.0, 798154610.75, 824094635.6, 850877711.26, 878531236.88], 'Добавленная стоимость, тыс. руб.': [191057445.0, 221584931.0, 226103515.0, 230716026.71, 235422633.65, 240225255.38, 245125850.59]}
        >>> get_temp_rosta(data)
        {'ВТП Казани, тыс. руб.': 103.25, 'Добавленная стоимость, тыс. руб.': 102.04}
    '''

    return {key: round((DS_otrasli[key][-1] / DS_otrasli[key][-2]) * 100, 2) if len(DS_otrasli[key]) > 1 else 0 for key in list(DS_otrasli.keys())}


def get_k(DS_otrasli):
    '''
    Возвращает значение К для вычисления прогнозируемых значений
    На вход функции подается словарь со значениями добавленной стоимости по отраслям 
    На выход функции подается словарь с коэффициетнтами К для дальнейших подсчетов 

    >>> CORRECT_TITLES = ['ВТП Казани, тыс. руб.', 'Добавленная стоимость, тыс. руб.', 'ДС Обрабатывающие производства, тыс. руб.', 'ДС Строительство, тыс. руб.', 'ДС Транспортировка и хранение, тыс. руб.', 'ДС Оптовая и розничная торговля, тыс. руб.', 'ДС Деятельность в области информатизации и связи, тыс. руб.']
    >>> DS_otrasli = {'ВТП Казани, тыс. руб.': [674739900.0, 748722900.0, 773031100.0, 798154610.75, 824094635.6, 850877711.26, 878531236.88], 'Добавленная стоимость, тыс. руб.': [191057445.0, 221584931.0, 226103515.0, 230716026.71, 235422633.65, 240225255.38, 245125850.59], 'ДС Обрабатывающие производства, тыс. руб.': [75373868.0, 80169194.0, 76066267.0, 78783598.344, 79958359.172, 79850189.785], 'ДС Строительство, тыс. руб.': [8376613.0, 9553660.0, 8194075.0, 8888142.345, 9065600.551, 8888954.769], 'ДС Транспортировка и хранение, тыс. руб.': [26272378.0, 35407329.0, 2987596.0, 22157415.832, 20690880.642, 15470346.997], 'ДС Оптовая и розничная торговля, тыс. руб.': [19963824.0, 25665497.0, 36181429.0, 27718999.183, 30453953.188, 32137036.174], 'ДС Деятельность в области информатизации и связи, тыс. руб.': [6646511.0, 9461344.0, 10263705.0, 8946153.853, 9756847.568, 9857412.822]}
    >>> get_k(DS_otrasli)
    {'ДС Обрабатывающие производства, тыс. руб.': 0.33108225106707384, 'ДС Строительство, тыс. руб.': 0.037251573626981824, 'ДС Транспортировка и хранение, тыс. руб.': 0.0811202094611359, 'ДС Оптовая и розничная торговля, тыс. руб.': 0.1252060518767167, 'ДС Деятельность в области информатизации и связи, тыс. руб.': 0.03960983031372719}
     '''
    DS = DS_otrasli[CORRECT_TITLES[1]]

    return {otrasl: (DS_otrasli[otrasl][-3] / DS[-3] + DS_otrasli[otrasl][-2] / DS[-2] + DS_otrasli[otrasl][-1] / DS[-1]) / 3 for otrasl in list(DS_otrasli.keys())[2:]}


def get_analyze(data):
    '''
        Возвращает оценку влияния прироста добавленной стоимости
        На вход функции подается словарь со значениями таблицы
        На выход функции подается словарь с данными по оценке прироста добавленной стоимости

        >>> data = {'ВТП Казани, тыс. руб.': [674739900.0, 748722900.0, 773031100.0, 798154610.75, 824094635.6, 850877711.26, 878531236.88], 'Добавленная стоимость, тыс. руб.': [191057445.0, 221584931.0, 226103515.0, 230716026.71, 235422633.65, 240225255.38, 245125850.59], 'ДС Обрабатывающие производства, тыс. руб.': [75373868.0, 80169194.0, 76066267.0, 78783598.344, 79958359.172, 79850189.785, 81156818.408], 'ДС Строительство, тыс. руб.': [8376613.0, 9553660.0, 8194075.0, 8888142.345, 9065600.551, 8888954.769, 9131323.671], 'ДС Транспортировка и хранение, тыс. руб.': [26272378.0, 35407329.0, 2987596.0, 22157415.832, 20690880.642, 15470346.997, 19884660.344], 'ДС Оптовая и розничная торговля, тыс. руб.': [19963824.0, 25665497.0, 36181429.0, 27718999.183, 30453953.188, 32137036.174, 30691239.965], 'ДС Деятельность в области информатизации и связи, тыс. руб.': [6646511.0, 9461344.0, 10263705.0, 8946153.853, 9756847.568, 9857412.822, 9709393.347]}
        >>> get_analyze(data)
        {'Добавленная стоимость, тыс. руб.': [4.524, 0.604, 0.597, 0.59, 0.583, 0.576], 'ДС Обрабатывающие производства, тыс. руб.': [0.711, -0.548, 0.352, 0.147, -0.013, 0.154], 'ДС Строительство, тыс. руб.': [0.174, -0.182, 0.09, 0.022, -0.021, 0.028], 'ДС Транспортировка и хранение, тыс. руб.': [1.354, -4.33, 2.48, -0.184, -0.633, 0.519], 'ДС Оптовая и розничная торговля, тыс. руб.': [0.845, 1.405, -1.095, 0.343, 0.204, -0.17], 'ДС Деятельность в области информатизации и связи, тыс. руб.': [0.417, 0.107, -0.17, 0.102, 0.012, -0.017]}
    '''
    delta_VDST = {}

    for i in range(len(data[CORRECT_TITLES[0]]) - 1):
        for otrasl in list(data.keys())[1:]:
            delta_DS = data[otrasl][i + 1] - data[otrasl][i]
            delta_VDST_val = delta_DS / data[CORRECT_TITLES[0]][i] * 100
            delta_VDST[otrasl] = delta_VDST.get(
                otrasl, []) + [round(delta_VDST_val, 3)]

    return delta_VDST



def get_data_excel_format(delta_VDST, DS):
    '''
        Возвращает список строк с необходмым заполнением excel документа данными о ДС
        На вход функции подаются словарь с данными по оценке прироста добавленной  стоимости и добавленная стоимость 
        На выход подается список строк с необходмым заполнением excel документа данными о ДС

        >>> delta_VDST = {'Добавленная стоимость, тыс. руб.': [4.524, 0.604, 0.597, 0.59, 0.583, 0.576], 'ДС Обрабатывающие производства, тыс. руб.': [0.711, -0.548, 0.352, 0.147, -0.013, 0.154], 'ДС Строительство, тыс. руб.': [0.174, -0.182, 0.09, 0.022, -0.021, 0.028], 'ДС Транспортировка и хранение, тыс. руб.': [1.354, -4.33, 2.48, -0.184, -0.633, 0.519], 'ДС Оптовая и розничная торговля, тыс. руб.': [0.845, 1.405, -1.095, 0.343, 0.204, -0.17], 'ДС Деятельность в области информатизации и связи, тыс. руб.': [0.417, 0.107, -0.17, 0.102, 0.012, -0.017]}
        >>> DS = {'ВТП Казани, тыс. руб.': [674739900.0, 748722900.0, 773031100.0, 798154610.75, 824094635.6, 850877711.26, 878531236.88], 'Добавленная стоимость, тыс. руб.': [191057445.0, 221584931.0, 226103515.0, 230716026.71, 235422633.65, 240225255.38, 245125850.59], 'ДС Обрабатывающие производства, тыс. руб.': [75373868.0, 80169194.0, 76066267.0, 78783598.344, 79958359.172, 79850189.785, 81156818.408], 'ДС Строительство, тыс. руб.': [8376613.0, 9553660.0, 8194075.0, 8888142.345, 9065600.551, 8888954.769, 9131323.671], 'ДС Транспортировка и хранение, тыс. руб.': [26272378.0, 35407329.0, 2987596.0, 22157415.832, 20690880.642, 15470346.997, 19884660.344], 'ДС Оптовая и розничная торговля, тыс. руб.': [19963824.0, 25665497.0, 36181429.0, 27718999.183, 30453953.188, 32137036.174, 30691239.965], 'ДС Деятельность в области информатизации и связи, тыс. руб.': [6646511.0, 9461344.0, 10263705.0, 8946153.853, 9756847.568, 9857412.822, 9709393.347]}
        >>> get_data_excel_format(delta_VDST, DS)
        [['ВТП Казани, тыс. руб.', '', 674739900.0, 748722900.0, 773031100.0, 798154610.75, 824094635.6, 850877711.26, 878531236.88], ['Добавленная стоимость, тыс. руб.', '', 191057445.0, 221584931.0, 226103515.0, 230716026.71, 235422633.65, 240225255.38, 245125850.59], ['Оценка влияния прироста добавленной стоимости на ВТП', '', '-', 4.524, 0.604, 0.597, 0.59, 0.583, 0.576], ['Обрабатывающие производства, тыс. руб.', 'ДС', 75373868.0, 80169194.0, 76066267.0, 78783598.344, 79958359.172, 79850189.785, 81156818.408], ['', 'Впериод', '-', 0.711, -0.548, 0.352, 0.147, -0.013, 0.154], ['Строительство, тыс. руб.', 'ДС', 8376613.0, 9553660.0, 8194075.0, 8888142.345, 9065600.551, 8888954.769, 9131323.671], ['', 'Впериод', '-', 0.174, -0.182, 0.09, 0.022, -0.021, 0.028], ['Транспортировка и хранение, тыс. руб.', 'ДС', 26272378.0, 35407329.0, 2987596.0, 22157415.832, 20690880.642, 15470346.997, 19884660.344], ['', 'Впериод', '-', 1.354, -4.33, 2.48, -0.184, -0.633, 0.519], ['Оптовая и розничная торговля, тыс. руб.', 'ДС', 19963824.0, 25665497.0, 36181429.0, 27718999.183, 30453953.188, 32137036.174, 30691239.965], ['', 'Впериод', '-', 0.845, 1.405, -1.095, 0.343, 0.204, -0.17], ['Деятельность в области информатизации и связи, тыс. руб.', 'ДС', 6646511.0, 9461344.0, 10263705.0, 8946153.853, 9756847.568, 9857412.822, 9709393.347], ['', 'Впериод', '-', 0.417, 0.107, -0.17, 0.102, 0.012, -0.017]]
    '''
    rows = [
        [CORRECT_TITLES[0], ''] + DS[CORRECT_TITLES[0]],
        [CORRECT_TITLES[1], ''] + DS[CORRECT_TITLES[1]],
        ['Оценка влияния прироста добавленной стоимости на ВТП',
            '', '-'] + delta_VDST[CORRECT_TITLES[1]],
    ]

    for otrasl in CORRECT_TITLES[2:]:
        rows.append([otrasl[3:], 'ДС'] + DS[otrasl])
        rows.append(['', 'Впериод', '-'] + delta_VDST[otrasl])

    return rows


def cell_merge(ws):
    '''
        Объединяет необходимые строки в excel документе согласно необходимому формату
        На вход функции поступает значение с названием страницы файла типа excel, хранящее в себе таблицу
    '''
    for i in range(2, 5):
        ws.merge_cells(f'A{i}:B{i}')

    for i in range(5, 14, 2):
        ws.merge_cells(f'A{i}:A{i + 1}')


def save_to_excel(start_year, DS, delta_VDST):
    '''
        Сохраняет данные в excel документ
        На вход функции поступают: значение начала отчетного периода(год), словарь с добавленной стоимостью и словарь с данными по оценке прироста добавленной  стоимости
    '''
    wb = openpyxl.Workbook()
    ws = wb.active

    for i, value in enumerate(['', ''] + [start_year +
                          i for i in range(len(DS[CORRECT_TITLES[0]]))], 1):
        ws.cell(row=1, column=i).value = value

    for i, row in enumerate(get_data_excel_format(delta_VDST, DS), 2):
        ws.cell(row=i, column=1)
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j)
            cell.value = value
            cell.alignment = Alignment(wrapText=True)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10

    cell_merge(ws)
    wb.save('excel.xlsx')


def get_prognos(data):
    '''
        Возвращает прогнозные значения ДС
        На вход функции подается словарь со значениями таблицы
        На выход функции подается преобразованный словарь со значениями таблицы

        >>> data = {'ВТП Казани, тыс. руб.': [674739900.0, 748722900.0, 773031100.0], 'Добавленная стоимость, тыс. руб.': [191057445.0, 221584931.0, 226103515.0], 'ДС Обрабатывающие производства, тыс. руб.': [75373868.0, 80169194.0, 76066267.0], 'ДС Строительство, тыс. руб.': [8376613.0, 9553660.0, 8194075.0], 'ДС Транспортировка и хранение, тыс. руб.': [26272378.0, 35407329.0, 2987596.0], 'ДС Оптовая и розничная торговля, тыс. руб.': [19963824.0, 25665497.0, 36181429.0], 'ДС Деятельность в области информатизации и связи, тыс. руб.': [6646511.0, 9461344.0, 10263705.0]}
        >>> get_prognos(data)
        {'ВТП Казани, тыс. руб.': [674739900.0, 748722900.0, 773031100.0, 798154610.75, 824094635.6, 850877711.26, 878531236.88], 'Добавленная стоимость, тыс. руб.': [191057445.0, 221584931.0, 226103515.0, 230716026.71, 235422633.65, 240225255.38, 245125850.59], 'ДС Обрабатывающие производства, тыс. руб.': [75373868.0, 80169194.0, 76066267.0, 78783598.344, 79958359.172, 79850189.785, 81156818.408], 'ДС Строительство, тыс. руб.': [8376613.0, 9553660.0, 8194075.0, 8888142.345, 9065600.551, 8888954.769, 9131323.671], 'ДС Транспортировка и хранение, тыс. руб.': [26272378.0, 35407329.0, 2987596.0, 22157415.832, 20690880.642, 15470346.997, 19884660.344], 'ДС Оптовая и розничная торговля, тыс. руб.': [19963824.0, 25665497.0, 36181429.0, 27718999.183, 30453953.188, 32137036.174, 30691239.965], 'ДС Деятельность в области информатизации и связи, тыс. руб.': [6646511.0, 9461344.0, 10263705.0, 8946153.853, 9756847.568, 9857412.822, 9709393.347]}
    '''
    temp_rosta = get_temp_rosta(data)

    for i in range(years_count):
        # VTP(prognoz)
        data[CORRECT_TITLES[0]].append(
            round(data[CORRECT_TITLES[0]][-1] * (temp_rosta[CORRECT_TITLES[0]] / 100), 2))

        # DS(prognoz)
        data[CORRECT_TITLES[1]].append(
            round(data[CORRECT_TITLES[1]][-1] * (temp_rosta[CORRECT_TITLES[1]] / 100), 2))
        K = get_k(data)

        for otrasl in list(data.keys())[2:]:
            data[otrasl].append(
                round((K[otrasl] * data[CORRECT_TITLES[1]][-1]), 3))

    return data


def get_axis_x(data_df: pd.DataFrame,start_year: int) -> List:
    '''
    Функция формирует список со значениями(axis_x) для оси Х(горизонтальной)
    На вход функции подается Data Frame со значениями бюджета и категориями,
    а также год начала отчетного периода
    На выход функция подает готовые значения для оси Х (горизонтальная) в виде отчетного периода  в годах 

    >>> dict = {'ВТП Казани, тыс. руб.': [674739900.0, 748722900.0, 773031100.0, 798154610.75, 824094635.6, 850877711.26, 878531236.88], 'Добавленная стоимость, тыс. руб.': [191057445.0, 221584931.0, 226103515.0, 230716026.71, 235422633.65, 240225255.38, 245125850.59], 'ДС Обрабатывающие производства, тыс. руб.': [75373868.0, 80169194.0, 76066267.0, 78783598.344, 79958359.172, 79850189.785, 81156818.408], 'ДС Строительство, тыс. руб.': [8376613.0, 9553660.0, 8194075.0, 8888142.345, 9065600.551, 8888954.769, 9131323.671], 'ДС Транспортировка и хранение, тыс. руб.': [26272378.0, 35407329.0, 2987596.0, 22157415.832, 20690880.642, 15470346.997, 19884660.344], 'ДС Оптовая и розничная торговля, тыс. руб.': [19963824.0, 25665497.0, 36181429.0, 27718999.183, 30453953.188, 32137036.174, 30691239.965], 'ДС Деятельность в области информатизации и связи, тыс. руб.': [6646511.0, 9461344.0, 10263705.0, 8946153.853, 9756847.568, 9857412.822, 9709393.347]}
    >>> start_year = 2017
    >>> data_df = pd.DataFrame(dict) 
    >>> get_axis_x(data_df,start_year)
    ['2017', '2018', '2019', '2020', '2021', '2022', '2023']
    '''
    axis_x = [str(i) for i in range(start_year, start_year+len(data_df))]
    data_df["Период"] = axis_x 

    return axis_x


def make_histogram(data_df: pd.DataFrame, axis_x: List):
    '''
    Функция формирует значения для построения графиков
    и строит графики, сохраняя их с заданным расширением 
    На вход функции подается Data Frame объект и значения для оси Х (горизонтальной)
    '''

    for key in data_df.keys():
        if key == 'Период': continue
        fig = px.histogram(data_df, x="Период", y=key, color=["c"]*3+["r"]*4, title=key)
        fig.update_layout(
        xaxis_title_text = 'Период (*красный - прогнозируемый год)',
        yaxis_title_text = key,
        bargap = 0.1,
        template = 'simple_white') 
        fig.write_html(key + '.html',
               include_plotlyjs='cdn',
               full_html=False,
               include_mathjax='cdn')             
    

if check_params(document_inner) and not is_empty(document_inner):
    data_df = document_inner.set_index('Наименование').T
    data = {item[0]: list(item[1].values())
            for item in data_df.to_dict().items()}
    data = get_prognos(data)
    save_to_excel(int((document_inner.columns)[
                    1]), data, get_analyze(data))


data_df = pd.DataFrame(data)
start_year = int((document_inner.columns)[1])
axis_x = get_axis_x(data_df, start_year)
make_histogram(data_df, axis_x)